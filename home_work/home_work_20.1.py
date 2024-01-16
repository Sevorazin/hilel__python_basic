import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Читання даних з CSV-файлу
csv_file_path = r'D:\python_project\hilel_python_basic\lesson_11\home_work_19\output_data.csv'
data = []
with open(csv_file_path, 'r', newline='', encoding='utf-8') as csv_file:
    csv_reader = csv.DictReader(csv_file)
    for row in csv_reader:
        data.append(row)

# Видалення стовпця "Вік"
for row in data:
    del row['Вік']

# Запис даних у Excel-файл
excel_file_path = 'output_data.xlsx'
wb = Workbook()
ws = wb.active

# Записує назви стовпців
col_num = 1
for header in data[0].keys():
    col_letter = get_column_letter(col_num)
    ws[f'{col_letter}1'] = header
    col_num += 1

# Записує дані у відповідних рядках
for row_num, row_data in enumerate(data, 2):
    col_num = 1
    for value in row_data.values():
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}{row_num}'] = value
        col_num += 1

wb.save(excel_file_path)

print(f"Дані були успішно записані у файл '{excel_file_path}'")
