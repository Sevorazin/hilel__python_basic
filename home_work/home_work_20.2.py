from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

excel_file_path = 'output_data.xlsx'
wb = load_workbook(excel_file_path)
ws = wb.active

transposed_wb = Workbook()
transposed_ws = transposed_wb.active


iteration_count = 0

for col_num in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_num)

    for row_num in range(1, ws.max_row + 1):
        transposed_ws[f'{get_column_letter(row_num)}{col_num + 1}'] = ws[f'{col_letter}{row_num}'].value
        iteration_count += 1

repeat_count = (int(iteration_count) - 3) // 3

for col_num in range(1, repeat_count + 1):
    transposed_ws[f'{get_column_letter(col_num + 1)}1'] = f'персон_{col_num}'


transposed_excel_file_path = 'transposed_output_data.xlsx'
transposed_wb.save(transposed_excel_file_path)

print(f"Данные были успешно транспонированы и записаны в файл '{transposed_excel_file_path}'")
