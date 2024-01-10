import openpyxl


def delete_person():
    try:
        wb = openpyxl.load_workbook('database.xlsx')
        ws = wb.active
    except FileNotFoundError:
        print("База даних не знайдена.")
        return

    headers = [cell.value for cell in ws[1]]

    print("Виберіть параметр для пошуку:")
    for i, header in enumerate(headers, start=1):
        print(f"{i}. {header}")

    search_choice = int(input("Ваш вибір: "))

    if 1 <= search_choice <= len(headers):
        search_column = headers[search_choice - 1]
    else:
        print("Невірний вибір параметра.")
        return

    search_value = input(f"Введіть значення для параметру '{search_column}': ")

    results = []
    for i, row in enumerate(
            ws.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=1
    ):
        person = dict(zip(headers, row))
        if str(person[search_column]) == search_value:
            results.append((i, person))

    if not results:
        print("Людина не знайдена за обраним параметром.")
        return

    print("\nЗнайдені люди:")
    print("=" * 151)
    print("""
| {:<5} | {:<15} | {:<15} | {:<15} | {:<15} | {:<15} | {:<15} | {:<31} |
""".format(
        '№',
        *headers
    ))
    print("=" * 151)
    for i, (row_num, result) in enumerate(results, start=1):
        print("""
| {:<5} | {:<15} | {:<15} | {:<15} | {:<15} | {:<15} | {:<15} | {:<31} |
""".format(
            i,
            result.get('Ім\'я', '') or '',
            result.get('Прізвище', '') or '',
            result.get('По-батькові', '') or '',
            result.get('Стать', '') or '',
            result.get('Дата народження', '') or '',
            result.get('Дата смерті', '') or 'Відсутня',
            result.get('Вік', '') or ''
        ))
    print("=" * 151)

    delete_choice = input("""
Введіть номер людини,
яку ви хочете видалити
(розділіть комами для видалення декількох): """
                          )

    try:
        delete_indices = [int(index) for index in delete_choice.split(",")]
    except ValueError:
        print("Неправильний формат введення.")
        return

    for index in sorted(delete_indices, reverse=True):
        if 1 <= index <= len(results):
            ws.delete_rows(results[index - 1][0] + 1)
        else:
            print(f"Невірний номер людини: {index}")

    wb.save('database.xlsx')
    print("Людина(и) видалена(и) з бази даних.")


if __name__ == "__main__":
    delete_person()
