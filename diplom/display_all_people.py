from display_utils import display_person
import openpyxl


def display_all_people():
    try:
        wb = openpyxl.load_workbook('database.xlsx')
        ws = wb.active
    except FileNotFoundError:
        print("База даних не знайдена.")
        return

    headers = [cell.value for cell in ws[1]]

    print("\nУсі люди в базі даних:")
    print("=" * 143)
    print("""
| {:<15} | {:<15} | {:<15} | {:<15} | {:<15} | {:<15} | {:<31} |
""".format(
        *headers
    ))
    print("=" * 143)
    for row in ws.iter_rows(min_row=2, values_only=True):
        person = dict(zip(headers, row))
        display_person(person)


if __name__ == "__main__":
    display_all_people()
