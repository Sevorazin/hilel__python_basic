from display_utils import display_person
import openpyxl


def search_for_a_person():
    previous_choice = None

    while True:
        print("Виберіть параметр для пошуку:")
        print("1. Ім'я")
        print("2. Прізвище")
        print("3. По-батькові")
        print("4. Стать")
        print("5. Дата народження")
        print("6. Дата смерті")
        print("7. Вік")

        choice = input("Ваш вибір: ")

        if choice.isdigit() and 1 <= int(choice) <= 7:
            break
        else:
            print("Невірний ввід. Будь ласка, введіть число від 1 до 7.")

    try:
        wb = openpyxl.load_workbook('database.xlsx')
        ws = wb.active
    except FileNotFoundError:
        print("База даних не знайдена.")
        return

    if previous_choice is not None and choice != previous_choice:
        print("Параметр змінився. Починаю пошук з початку.")

    headers = [cell.value for cell in ws[1]]

    search_value = input(
        f"""
Введіть значення для параметру '{headers[int(choice) - 1]}':
"""
                         )

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        person = dict(zip(headers, row))
        if str(person[headers[int(choice) - 1]]) == search_value:
            results.append(person)

    if results:
        print("\nРезультати пошуку:")
        print("=" * 143)
        print("""
| {:<15} | {:<15} | {:<15} | {:<15} | {:<15} | {:<15} | {:<31} |
""".format(
            *headers
        ))
        print("=" * 143)
        for result in results:
            display_person(result)
    else:
        print("Людина не знайдена за обраним параметром.")
        choice_retry = input("""
1. Спробувати інший параметр
2. Повернутися в меню
Ваш вибір: """
                             )

        if choice_retry == '1':
            search_for_a_person()
        elif choice_retry == '2':
            return
        else:
            print("Невірний ввід. Повертаюсь в меню.")
            return


if __name__ == "__main__":
    search_for_a_person()
